import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = "/home/claude/bank_dashboard_project/data"
OUT = "/home/claude/bank_dashboard_project/data/bank_financial_data.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F2937", end_color="1F2937")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=10)        # blue = hardcoded input
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)      # black = formula
LINK_FONT = Font(name=FONT_NAME, color="008000", size=10)         # green = cross-sheet link
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="6B7280")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(name):
    with open(f"{DATA}/{name}") as f:
        return list(csv.DictReader(f))


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------------------------------------------------------- Banks ----
banks = read_csv("banks.csv")
ws = wb.active
ws.title = "Banks"
headers = ["BankID", "BankName", "Ticker", "Country", "Region", "HQCity", "Latitude", "Longitude", "ReportingCurrency"]
ws.append(headers)
style_header(ws, 1, len(headers))
for b in banks:
    row = [int(b["BankID"]), b["BankName"], b["Ticker"], b["Country"], b["Region"], b["HQCity"],
           float(b["Latitude"]), float(b["Longitude"]), b["ReportingCurrency"]]
    ws.append(row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).font = INPUT_FONT
        ws.cell(row=r, column=c).border = BORDER
autofit(ws, [8, 20, 8, 16, 10, 14, 10, 10, 16])
ws.freeze_panes = "A2"

# ------------------------------------------------------------- FX_Rates ----
fx = read_csv("fx_rates.csv")
ws2 = wb.create_sheet("FX_Rates")
headers = ["Year", "EUR_to_USD", "GBP_to_USD"]
ws2.append(headers)
style_header(ws2, 1, len(headers))
for r in fx:
    ws2.append([int(r["Year"]), float(r["EUR_to_USD"]), float(r["GBP_to_USD"])])
    rr = ws2.max_row
    for c in range(1, 4):
        ws2.cell(row=rr, column=c).font = INPUT_FONT
        ws2.cell(row=rr, column=c).border = BORDER
ws2.cell(row=1, column=4, value="Source: approximate annual average market rates (ECB / Federal Reserve historical series). Verify exact figures if precision is required for grading.").font = NOTE_FONT
autofit(ws2, [8, 14, 14, 90])

# --------------------------------------------------------- Financials_Raw --
fin = read_csv("financials.csv")
ws3 = wb.create_sheet("Financials_Raw")
headers = ["BankID", "Year", "Currency", "Revenue_M", "NetIncome_M", "TotalAssets_M", "ROE_Pct", "CET1_Pct", "Employees"]
ws3.append(headers)
style_header(ws3, 1, len(headers))
for r in fin:
    ws3.append([int(r["BankID"]), int(r["Year"]), r["Currency"], float(r["Revenue_M"]), float(r["NetIncome_M"]),
                float(r["TotalAssets_M"]), float(r["ROE_Pct"]), float(r["CET1_Pct"]), int(r["Employees"])])
    rr = ws3.max_row
    for c in range(1, len(headers) + 1):
        ws3.cell(row=rr, column=c).font = INPUT_FONT
        ws3.cell(row=rr, column=c).border = BORDER
ws3.cell(row=1, column=10, value="Source: company annual reports / earnings releases (10-K, 20-F, FY results press releases), 2021-2025. Figures in millions of native currency.").font = NOTE_FONT
autofit(ws3, [8, 8, 10, 12, 12, 14, 10, 10, 12, 90])
ws3.freeze_panes = "A2"
n_fin_rows = len(fin)

# ----------------------------------------------------------- Financials_USD (FORMULAS) --
ws4 = wb.create_sheet("Financials_USD")
headers = ["BankID", "BankName", "Year", "Currency", "Revenue_Native_M", "NetIncome_Native_M",
           "FX_Rate_to_USD", "Revenue_USD_M", "NetIncome_USD_M", "NetMargin_Pct", "ROE_Pct", "CET1_Pct"]
ws4.append(headers)
style_header(ws4, 1, len(headers))

for i in range(2, n_fin_rows + 2):
    r = i  # row in this sheet == row in Financials_Raw (same order, header offset matches)
    ws4.cell(row=r, column=1, value=f"=Financials_Raw!A{r}").font = LINK_FONT
    ws4.cell(row=r, column=2, value=f"=VLOOKUP(A{r},Banks!$A:$B,2,FALSE)").font = LINK_FONT
    ws4.cell(row=r, column=3, value=f"=Financials_Raw!B{r}").font = LINK_FONT
    ws4.cell(row=r, column=4, value=f"=Financials_Raw!C{r}").font = LINK_FONT
    ws4.cell(row=r, column=5, value=f"=Financials_Raw!D{r}").font = LINK_FONT
    ws4.cell(row=r, column=6, value=f"=Financials_Raw!E{r}").font = LINK_FONT
    # FX rate: 1 for USD, else lookup EUR or GBP column in FX_Rates by Year
    ws4.cell(row=r, column=7,
             value=f'=IF(D{r}="USD",1,IF(D{r}="EUR",VLOOKUP(C{r},FX_Rates!$A:$C,2,FALSE),VLOOKUP(C{r},FX_Rates!$A:$C,3,FALSE)))'
             ).font = FORMULA_FONT
    ws4.cell(row=r, column=8, value=f"=E{r}*G{r}").font = FORMULA_FONT
    ws4.cell(row=r, column=9, value=f"=F{r}*G{r}").font = FORMULA_FONT
    ws4.cell(row=r, column=10, value=f"=IF(H{r}=0,0,I{r}/H{r}*100)").font = FORMULA_FONT
    ws4.cell(row=r, column=11, value=f"=Financials_Raw!G{r}").font = LINK_FONT
    ws4.cell(row=r, column=12, value=f"=Financials_Raw!H{r}").font = LINK_FONT
    for c in range(1, len(headers) + 1):
        ws4.cell(row=r, column=c).border = BORDER
        if c in (8, 9):
            ws4.cell(row=r, column=c).number_format = "#,##0"
        if c == 10:
            ws4.cell(row=r, column=c).number_format = "0.0"

ws4.cell(row=1, column=14, value="This sheet recreates the USD conversion using live Excel formulas (VLOOKUP into Banks / FX_Rates) so you can see the calculation step-by-step. The Java program performs the same conversion programmatically, plus extra metrics (YoY growth, Financial Health Score) — see the Processed_for_PowerBI sheet.").font = NOTE_FONT
autofit(ws4, [8, 20, 8, 10, 16, 16, 14, 14, 15, 13, 10, 10, 4, 100])
ws4.freeze_panes = "A2"

# ----------------------------------------------------- Processed_for_PowerBI --
proc = read_csv("processed_bank_data.csv")
ws5 = wb.create_sheet("Processed_for_PowerBI")
headers = list(proc[0].keys())
ws5.append(headers)
style_header(ws5, 1, len(headers))
numeric_cols = {"BankID", "Latitude", "Longitude", "Year", "Revenue_USD_M", "NetIncome_USD_M",
                 "TotalAssets_USD_M", "ROE_Pct", "CET1_Pct", "Employees", "NetMargin_Pct",
                 "RevenueGrowth_YoY_Pct", "NetIncomeGrowth_YoY_Pct", "FinancialHealthScore"}
for row in proc:
    vals = []
    for h in headers:
        v = row[h]
        if h in numeric_cols and v != "":
            v = float(v) if "." in v or h in ("Latitude", "Longitude", "RevenueGrowth_YoY_Pct", "NetIncomeGrowth_YoY_Pct", "FinancialHealthScore", "NetMargin_Pct") else int(v)
        vals.append(v)
    ws5.append(vals)
    rr = ws5.max_row
    for c in range(1, len(headers) + 1):
        ws5.cell(row=rr, column=c).font = FORMULA_FONT
        ws5.cell(row=rr, column=c).border = BORDER
ws5.cell(row=1, column=len(headers) + 2,
         value="This is the final, analysis-ready table generated by BankFinancialAnalyzer.java. Import THIS sheet (or the equivalent processed_bank_data.csv) into Power BI.").font = NOTE_FONT
autofit(ws5, [8, 20, 8, 16, 10, 12, 9, 12, 6, 14, 14, 15, 9, 9, 10, 11, 16, 18, 18])
ws5.freeze_panes = "A2"

# ------------------------------------------------------------------ ReadMe --
ws6 = wb.create_sheet("ReadMe")
ws6["A1"] = "Top 10 Global Banks — Financial Health Dashboard (2021–2025)"
ws6["A1"].font = TITLE_FONT
lines = [
    "",
    "RESEARCH QUESTION",
    "How does the overall financial health (profitability, efficiency, capital strength, and momentum) of the",
    "largest European banks compare with the largest American banks, and how has this evolved 2021-2025?",
    "",
    "SHEETS IN THIS WORKBOOK",
    "  Banks                  - dimension table: bank name, country, region, HQ coordinates (for the Power BI map)",
    "  FX_Rates                - approximate annual average EUR->USD and GBP->USD rates, used for currency normalization",
    "  Financials_Raw           - raw reported figures per bank/year, in NATIVE currency (as published by each bank)",
    "  Financials_USD          - same data converted to USD using live Excel formulas (VLOOKUP/IF) - inspect column G & H",
    "  Processed_for_PowerBI - final clean table (USD, +YoY growth, +Financial Health Score) produced by the Java program",
    "",
    "DATA SOURCES",
    "  Figures compiled from each bank's FY2021-FY2025 annual reports, 10-K / 20-F filings, and Q4/full-year earnings",
    "  press releases (JPMorgan Chase, Bank of America, Citigroup, Wells Fargo, Goldman Sachs, BNP Paribas, Banco",
    "  Santander, Deutsche Bank, Barclays, UBS Group). 2024-2025 figures were verified against company press releases;",
    "  earlier years are sourced from published annual figures and may be lightly rounded. For citation-grade academic",
    "  work, cross-check exact figures against the original 10-K/20-F/annual report filings before submission.",
    "",
    "RECOMMENDED PIPELINE",
    "  1. (This file) Review/adjust the raw data in Financials_Raw if you want to update or extend it.",
    "  2. Run the Java program (java/BankFinancialAnalyzer.java) to regenerate processed_bank_data.csv.",
    "  3. Run the Node.js script (js/validate_and_summarize.js) to check data quality and view the preview dashboard.",
    "  4. Open Power BI Desktop and follow docs/PowerBI_Guide.md to build the interactive dashboard.",
    "",
    "See README.md in the project root for the full step-by-step guide.",
]
for i, line in enumerate(lines, start=2):
    cell = ws6.cell(row=i, column=1, value=line)
    if line.isupper() and line != "":
        cell.font = Font(name=FONT_NAME, bold=True, size=11)
    else:
        cell.font = Font(name=FONT_NAME, size=10)
autofit(ws6, [115])

wb.move_sheet("ReadMe", offset=-(len(wb.sheetnames) - 1))  # put ReadMe first
wb.save(OUT)
print("Saved", OUT)
